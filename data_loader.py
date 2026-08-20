"""
Data loading layer for the HD/LOW comp sales forecast Streamlit app.
 
Reads directly from data/HD_LOW_Quarterly_Data.xlsx using cached (already
computed) cell values -- the workbook was fully recalculated with LibreOffice
headless (recalc.py) before being placed here, so pandas/openpyxl can read
the formula RESULTS without needing Excel or LibreOffice installed on
whatever machine (or Streamlit Cloud container) runs this app.
 
All tab layouts here (row/column positions) mirror exactly what the
build_*.py scripts in the main project wrote -- if that workbook's structure
changes, these constants need to change too.
"""
import pandas as pd
import openpyxl
import streamlit as st
 
DATA_FILE = "data/HD_LOW_Quarterly_Data.xlsx"
 
QTR_ROWS_HIST = {"HD": (5, 34), "LOW": (39, 68)}
QTR_ROWS_PROJ = {"HD": (35, 38), "LOW": (69, 72)}
 
QUARTERLY_COLUMNS = [
    "Company", "Fiscal Quarter", "Quarter Start", "Quarter End",
    "Mortgage Rate (%)", "Existing Home Sales (M, SAAR)", "New Home Sales (M, SAAR)",
    "CPI-U Y/Y Growth (%)", "Consumer Sentiment", "Building Permits (M, SAAR)",
    "Macro Data Notes", "Comp Sales % (US)", "NAICS444 Y/Y Growth",
    "LIRA Mapped Rate", "Comp/NAICS/LIRA Notes", "Period Type",
]
 
 
@st.cache_data
def load_quarterly_data() -> pd.DataFrame:
    """Full Quarterly Data tab (both companies, actual + projected rows), one row per fiscal quarter."""
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["Quarterly Data"]
    rows = []
    for company, (h_first, h_last) in QTR_ROWS_HIST.items():
        p_first, p_last = QTR_ROWS_PROJ[company]
        for r in list(range(h_first, h_last + 1)) + list(range(p_first, p_last + 1)):
            rows.append([ws.cell(row=r, column=c).value for c in range(1, 17)])
    df = pd.DataFrame(rows, columns=QUARTERLY_COLUMNS)
    df["Quarter Start"] = pd.to_datetime(df["Quarter Start"])
    df["Quarter End"] = pd.to_datetime(df["Quarter End"])
    df["Fiscal Year"] = df["Fiscal Quarter"].str.extract(r"FY(\d+)").astype(int)
    return df
 
 
@st.cache_data
def load_overview() -> pd.DataFrame:
    """The 8 point estimates (Overview tab, rows 7-14)."""
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["Overview"]
    rows = []
    for r in range(7, 15):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, 7)])
    return pd.DataFrame(rows, columns=[
        "Company", "Fiscal Quarter", "Predictor Used", "Predictor Value",
        "Point Estimate", "LIRA Projection (15% Blend)",
    ])
 
 
@st.cache_data
def load_macro_correlation() -> pd.DataFrame:
    """Macro vs Comp Correlation summary table, both companies, 5 indicators (Building Permits dropped)."""
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["Macro vs Comp Correlation"]
    rows = []
    for r in range(74, 80):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, 8)])
    for r in range(81, 87):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, 8)])
    df = pd.DataFrame(rows, columns=[
        "Company", "Macro Series", "Contemporaneous", "Leads 1Q", "Leads 2Q", "Leads 3Q", "Leads 4Q",
    ])
    return df[df["Macro Series"] != "Building Permits (M Units)"].reset_index(drop=True)
 
 
@st.cache_data
def load_naics_correlation() -> pd.DataFrame:
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["NAICS444 vs Comp Correlation"]
    rows = []
    for r in range(8, 13):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 7)]
        if vals[0]:
            rows.append(vals)
    return pd.DataFrame(rows, columns=[
        "Company", "Sample", "N", "Contemporaneous r", "Retail Leads 1Q r", "Retail Lags 1Q r",
    ])
 
 
@st.cache_data
def load_lira_series() -> pd.DataFrame:
    """Raw per-quarter LIRA table (rows 5-20): LIRA quarter, mapped fiscal quarter, period type,
    LIRA rate of change, and HD/LOW comp (blank for the 4 projected LIRA quarters)."""
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["LIRA vs Comp Correlation"]
    rows = []
    for r in range(5, 21):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, 8)])
    return pd.DataFrame(rows, columns=[
        "LIRA Quarter", "Fiscal Quarter", "Period Type", "LIRA 4Q Total ($B)",
        "LIRA Rate of Change", "HD Comp % (US)", "LOW Comp % (US)",
    ])
 
 
@st.cache_data
def load_lira_correlation() -> pd.DataFrame:
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["LIRA vs Comp Correlation"]
    rows = []
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 7)]
        if vals[0] in ("HD", "LOW"):
            rows.append(vals)
    return pd.DataFrame(rows, columns=[
        "Company", "Contemporaneous r", "Leads 1Q r", "Leads 2Q r", "Leads 3Q r", "N (Contemp.)",
    ])
 


